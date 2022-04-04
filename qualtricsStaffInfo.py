##
# qualtricsStaffInfo.py
# Paolo Reyes
# HPSD
#
# This file is responsible for the initial import and maintenance of staff data into the Qualtrics database. The script sources the
# staff data from an active directory export file that is ran on a daily basis to ensure the Qualtrics directory is consistently up to date.
#
# Methods:
#
# importADStaff()
#   Responsible for dividing the AD export list into chunks in order for each thread to import into Qualtrics.
#
# updateADStaff()
#   Responsible for dividing the AD export list and Qualtrics directory into chunks in order for each thread
#   to check for any updates in terms of contact details and new entries.
#
# getContactFromID(bearerToken, chunk, adList)  
#   Used by a single thread to loop through contact ids in it's unique chunk and calls function to gather their full details
# checkNewStaff(bearerToken, chunk, qExtRefs)  
#   Used by a single thread to loop through Qualtrics contacts in it's unique chunk to check if it exists in the AD export list
#
# formatContact(bearerToken, contactId, adList) 
#   API call to gather the full details from a specific contact id (embedded data is not included in List Directory Contacts API call).
#   Function then formats all this information in a consistent format to be easily compared with their entry in AD
#
# compareContactToAD(bearerToken, contactDetails, adList, contactId)  
#   Compares all the fields between a Qualtrics contact and their entry from AD export. Calls function to update the contact with new details
#   or delete the contact if not found in AD.
#
# getQualtricsBearer()  
#   Uses unique client id and secret key to generate a bearer token thats remains valid for one hour.
#
# getAllContacts(bearerToken)  
#   API call to retrieve directory contacts in bulk and returns a list of all their contact ids. This API call does not
#   return each contact's full list of information.
#
# createContact(bearerToken, staffDetails)  
#   API call to create a new Qualtrics contact based on the provided staff details. This function is also responsible for
#   handling cases where an entry from AD should not be created (test and guest accounts)
#
# updateContact(bearerToken, staffDetails, contactId)
#   API call to update a Qualtrics contact based on the provided staff details.
#
# deleteContact(bearerToken, contactId, staffDetails)  
#   API call to delete a Qualtrics contact from the specified directory.
#
# getContact(bearerToken, extRef)  
#   API call that checks if a provided employee id currently exists in the Qualtrics directory. This function is primarily used
#   to handle 504 gateway timeout errors as it determines whether a previous writing API call needs to be retried.




################################# IMPORT STATEMENTS #########################################

import requests
import json
import time
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import threading
from threading import Semaphore
import csv

################################# GLOBAL VARIABLES #########################################

logFile = open(str(time.strftime("%Y-%m-%d")) + " Log File.txt", "w") #log file created in same location as script
totalImported = 0
screenlock = Semaphore(value=1)

DATA_CENTER = "" #datacenter is found in your Account Settings and is 2-3 letters and a number
DIRECTORY = "" #directory id for which you are creating/updating contacts for

################################# IMPORT AD STAFF #########################################

def importADStaff():
    global totalImported
    global logFile
    allStaff = readADExport("PATH OF ACTIVE DIRECTORY STAFF EXPORT FILE")
    bearerToken = getQualtricsBearer()
    splitDump = list(split(allStaff, 8)) #run in parallel using 8 threads

    start_time = time.time()
    threads = []

    for chunk in splitDump:
        t = threading.Thread(target=loopThroughChunk, args=(bearerToken, chunk))
        threads.append(t)

    for t in threads:
        t.start()
    for t in threads:
        t.join()
    
    print("Total contacts imported:", totalImported, file=logFile)
    print("Function took", (time.time() - start_time)/60, "minutes to run", file=logFile)
    return

def loopThroughChunk(bearerToken, chunk):
    for details in chunk:
        createContact(bearerToken, details)

################################# MAINTAIN AD STAFF #########################################

def updateADStaff():
    adList = readADExportCSV("PATH OF ACTIVE DIRECTORY STAFF EXPORT FILE")
    bearerToken = getQualtricsBearer()

    qList = getAllContacts(bearerToken)
    allExtRefs = qList[0]
    allContactIds = qList[1]
    splitContactIds = list(split(allContactIds, 16)) #run parallel in 16 threads
    splitADList = list(split(adList, 16))

    start_time = time.time()
    threads = []
    threads2 = []

    for chunk in splitContactIds:
        t = threading.Thread(target= getContactFromID, args=(bearerToken, chunk, adList))
        threads.append(t)

    for t in threads:
        t.start()
    for t in threads:
        t.join()

    for chunk2 in splitADList:
        t2 = threading.Thread(target=checkNewStaff, args=(bearerToken, chunk2, allExtRefs))
        threads2.append(t2)

    for t2 in threads2:
        t2.start()
    for t2 in threads2:
        t2.join()

    print("Function took", (time.time() - start_time)/60, "minutes to run", file=logFile)
    return

def getContactFromID(bearerToken, chunk, adList):
    for contactId in chunk:
        formatContact(bearerToken, contactId, adList)
    return


def checkNewStaff(bearerToken, chunk, qExtRefs):
    for adEntry in chunk:
        adExtRef = str(adEntry[7])
        if adExtRef not in qExtRefs and adExtRef != '': #entry currently not in Qualtrics directory
            adFirstName = adEntry[2]
            adLastName = adEntry[1]
            adEmail = adEntry[4]
            adPhone = adEntry[6]
            adTitle = adEntry[5]
            adSite = adEntry[3]
            adDescription = adEntry[0]
            createContact(bearerToken, [adDescription, adLastName, adFirstName, adSite, adEmail, adTitle, adPhone, adExtRef])
    return


def formatContact(bearerToken, contactId, adList): #retrieves qualtrics contact and formats it accordingly to allow comparison with AD details
    baseURL = "https://{0}.qualtrics.com/API/v3/directories/".format(DATA_CENTER)
    auth = 'Bearer ' + bearerToken
    headers = {
        'Authorization': auth,
        'Content-Type': 'application/json'
    }
    test = baseURL + DIRECTORY + contactId
        
    initResponse = requests.get(baseURL + DIRECTORY + '/contacts/' + contactId, headers=headers)
    response = initResponse.json()
    screenlock.acquire()
    if response['meta']['httpStatus'][0:3] != '200': #api call returned error
        print("ERROR RETRIEVING CONTACT: RECEIVED", initResponse, "for", contactId, file=logFile)
        time.sleep(5)
        screenlock.release()
        formatContact(bearerToken, contactId, adList)
    else: #api call was successful
        screenlock.release()
        firstName = response['result']['firstName']
        lastName = response['result']['lastName']
        email = response['result']['email']
        phone = response['result']['phone']
        extRef = response['result']['extRef']
        embeddedData = response['result']['embeddedData']
        if 'Title' not in embeddedData:
            embeddedData['Title'] = ''
        if 'Site' not in embeddedData:
            embeddedData['Site'] = ''
        if 'Description' not in embeddedData:
            embeddedData['Description'] = ''
        contactDetails = [firstName, lastName, email, phone, extRef, embeddedData]
        contactDetails = ["" if value == None else value for value in contactDetails]
        compareContactToAD(bearerToken, contactDetails, adList, contactId)
    return 


#this will check if there are any value changes for staff, also checks if there is any qualtrics contacts that need to be deleted
def compareContactToAD(bearerToken, contactDetails, adList, contactId):
    qFirstName = contactDetails[0]
    qLastName = contactDetails[1]
    qEmail = contactDetails[2]
    qPhone = contactDetails[3]
    qExtRef = contactDetails[4]
    qTitle = contactDetails[5]['Title']
    qSite = contactDetails[5]['Site']
    qDescription = contactDetails[5]['Description']

    found = False

    for adEntry in adList:
        adExtRef = str(adEntry[7])
        if qExtRef == adExtRef: #found entry in qualtrics
            found = True
            adFirstName = adEntry[2]
            adLastName = adEntry[1]
            adEmail = adEntry[4]
            adPhone = adEntry[6]
            adTitle = adEntry[5]
            adSite = adEntry[3]
            adDescription = adEntry[0]
            newDetails = [adFirstName, adLastName, adEmail, adPhone, adExtRef, adTitle, adSite, adDescription]
            #check for changes between Qualtrics and AD details
            if adFirstName != qFirstName or adLastName != qLastName or adEmail != qEmail or adPhone != qPhone or adTitle != qTitle or adSite != qSite or adDescription != qDescription:
                updateContact(bearerToken, newDetails, contactId)
            else: #no changes required
                screenlock.acquire()
                print("NOTICE: No update needed for", qFirstName, qLastName, file=logFile)
                screenlock.release()
                return
    if found == False: #contact was not found in AD list, delete entry from Qualtrics
        deleteContact(bearerToken, contactId, contactDetails)
        return
    return


################################# QUALTRICS API CALLS #########################################

def getQualtricsBearer():
    #create the Base64 encoded basic authorization string
    clientId="UNIQUE OAUTH CLIENT ID"
    clientSecret="UNIQUE OAUTH CLIENT SECRET KEY"

    baseURL = "https://{0}.qualtrics.com/oauth2/token".format(DATA_CENTER) 
    data = {'grant_type': 'client_credentials','scope': 'manage:all'}
    
    r = requests.post(baseURL, auth=(clientId, clientSecret), data=data)
    return r.json()['access_token']


def getAllContacts(bearerToken):
    allExtRefs = []
    allContactIds = []
    baseURL = "https://{0}.qualtrics.com/API/v3/directories/".format(DATA_CENTER)
    auth = 'Bearer ' + bearerToken
    headers = {
      'Authorization': auth,
      'Content-Type': 'application/json'
    }
    param = {'useNewPaginationScheme': 'true'}

    response = requests.request("GET", baseURL + DIRECTORY + '/contacts?pageSize=100', headers=headers, params=param).json()
    if response['meta']['httpStatus'][0:3] == '200':
        while True:
            contactList = [entry['contactId'] for entry in response['result']['elements']]
            refList = [entry['extRef'] for entry in response['result']['elements']]
            allContactIds += contactList
            allExtRefs += refList
            #need to check for more pages as API call is only able to pull 100 contacts at most
            if response['result']['nextPage'] != None: 
                response = requests.request("GET", response['result']['nextPage'], headers=headers).json()
            else: #no more contacts left in directory
                break
    return [allExtRefs, allContactIds]


def createContact(bearerToken, staffDetails):
    baseURL = "https://{0}.qualtrics.com/API/v3/directories/".format(DATA_CENTER)
    auth = 'Bearer ' + bearerToken

    #if employee id exists(this check ensures any test/guest accounts do not get imported)
    if str(staffDetails[7]) != "":
        info = {
            "firstName": str(staffDetails[2]),
            "lastName": str(staffDetails[1]),
            "email": str(staffDetails[4]),
            "phone": str(staffDetails[6]),
            "extRef": str(staffDetails[7]),
            "embeddedData" : {"Title": staffDetails[5],
                              "Site": staffDetails[3],
                              "Description": staffDetails[0],
                              "Last Modified": time.strftime("%Y-%m-%d %H:%M")},
            "language": "",
            "unsubscribed": "False",
            }
        headers = {
          'Authorization': auth,
          'Content-Type': 'application/json',
        }

        response = requests.post(baseURL + DIRECTORY + '/contacts', headers=headers, json=info)
        responseCode = response.json()['meta']['httpStatus'][0:3]
        screenlock.acquire()
        if responseCode != '200': #unsuccessful API call
            print("ERROR: RECEIVED", response, "for", staffDetails[2], staffDetails[1], file=logFile)
            screenlock.release()
            if responseCode == '504': #handling Gateway Timeout error
                time.sleep(10)
                checkExistingContact = getContact(bearerToken, staffDetails[7])
                if not checkExistingContact: #contact does not exist
                    print("Attempting to retry import call for: ", staffDetails[2], staffDetails[1], file=logFile)
                    createContact(bearerToken, staffDetails)
                    return
            else: #generic API error, retry call
                time.sleep(5)
                createContact(bearerToken, staffDetails)
        else: #successful API call
            global totalImported
            totalImported += 1
            print(response, ": Successfully made contact for:", staffDetails[2], staffDetails[1], file=logFile)
            screenlock.release()
            return
    else: #no employee id exists, ignore entry
        screenlock.acquire()
        print("NOTICE: No Employee ID provided for", staffDetails[2], staffDetails[1], file=logFile)
        screenlock.release()
        return

def updateContact(bearerToken, staffDetails, contactId):
    baseURL = "https://{0}.qualtrics.com/API/v3/directories/".format(DATA_CENTER)
    auth = 'Bearer ' + bearerToken

    #contact details
    info = {
        "firstName": str(staffDetails[0]),
        "lastName": str(staffDetails[1]),
        "email": str(staffDetails[2]),
        "phone": str(staffDetails[3]),
        "extRef": str(staffDetails[4]),
        "embeddedData" : {"Title": staffDetails[5],
                            "Site": staffDetails[6],
                            "Description": staffDetails[7],
                            "Last Modified": time.strftime("%Y-%m-%d %H:%M")},
        "language": "",
        "unsubscribed": "False",
        }
    headers = {
        'Authorization': auth,
        'Content-Type': 'application/json',
    }

    response = requests.put(baseURL + DIRECTORY + '/contacts/' + contactId, headers=headers, json=info)
    
    test = response.json()
    screenlock.acquire()
    if response.json()['meta']['httpStatus'][0:3] != '200': #unsuccessful API call
        print("ERROR UPDATING CONTACT: RECEIVED", response, "for",  staffDetails[0], staffDetails[1], file=logFile)
        time.sleep(5) #wait 5 seconds before attempting again
        screenlock.release()
        updateContact(bearerToken, staffDetails, contactId)
    else: #successful API call
        print(response, ": Successfully updated details for:", staffDetails[0], staffDetails[1], file=logFile)
        screenlock.release()
        return

def deleteContact(bearerToken, contactId, staffDetails):
    baseURL = "https://{0}.qualtrics.com/API/v3/directories/".format(DATA_CENTER)
    auth = 'Bearer ' + bearerToken

    headers = {
        'Authorization': auth,
        'Content-Type': 'application/json'
    }

    response = requests.delete(baseURL + DIRECTORY + '/contacts/' + contactId, headers=headers)
    
    test = response.json()
    screenlock.acquire()
    if response.json()['meta']['httpStatus'][0:3] != '200': #unsuccessful API call
        print("ERROR DELETING CONTACT: RECEIVED", response, "for",  staffDetails[0], staffDetails[1], file=logFile)
        screenlock.release()
        if responseCode == '504': #handling Gateway Timeout error
            time.sleep(10)
            checkExistingContact = getContact(bearerToken, staffDetails[7])
            if checkExistingContact: #contact still in database
                print("Attempting to retry import call for: ", staffDetails[0], staffDetails[1], file=logFile)
                deleteContact(bearerToken, contactId, staffDetails)
                return
        else: #generic API error, retry call
            time.sleep(5)
            deleteContact(bearerToken, contactId, staffDetails)
    else: #successful API call
        print(response, ": Successfully deleted", staffDetails[0], staffDetails[1], "from directory", file=logFile)
        screenlock.release()
        return

def getContact(bearerToken, extRef):
    exists = False
    global screenlock
    baseURL = "https://{0}.qualtrics.com/API/v3/directories/".format(DATA_CENTER)
    auth = 'Bearer ' + bearerToken
    info = {
        "filter": {
            "filterType": "extRef",
            "comparison": "eq",
            "value": extRef     
            }
        }
    headers = {
      'Authorization': auth,
      'Content-Type': 'application/json',
    }

    response = requests.post(baseURL + DIRECTORY + '/contacts/search', headers=headers, json=info)
    test = response.json()
    screenlock.acquire()
    if response.json()['meta']['httpStatus'][0:3] != '200': #unsuccessful API call
        print("ERROR: RECEIVED", response, "WHEN ATTEMPTING TO GET CONTACT ID", extRef, file=logFile)
        time.sleep(5)
        screenlock.release()
        getContact(bearerToken, extRef)
        return
    else: #successful API call
        if len(response.json()['result']['elements']) > 0: #a contact will have atleast one completed field
            print("Contact", extRef, "exists in Qualtrics directory list", file=logFile)
            exists = True
        else: #empty array returned, contact does not exist
            print("Contact", extRef, "does not exists in Qualtrics directory list. Retrying import call...", file=logFile)
        screenlock.release()
        return exists


################################# HELPERS #########################################

def readADExport(path):
    wb = load_workbook(path)
    ws = wb['Sheet1']
    allCells = [['' if cell.value is None else cell.value for cell in row[3:]] for row in ws.iter_rows(min_row=2)]
    return allCells

def readADExportCSV(path):
    with open(path, 'r') as file:
        reader = csv.reader(file)
        allCells = [['' if cell is None else cell for cell in row[2:]] for row in reader]
    return allCells[1:]

def split(a, n):
    k, m = divmod(len(a), n)
    return (a[i*k+min(i, m):(i+1)*k+min(i+1, m)] for i in range(n))

def run():
    updateADStaff()
    global logFile
    logFile.close()

################################# CALLS #########################################

#importADStaff()
#updateADStaff()
run()
